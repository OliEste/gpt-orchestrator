"""
File conversion utilities for non-PDF files.
Converts Excel, DWG, and other file types into review artifacts (text, images, structured data).
"""
import os
import tempfile
import base64
import json
import csv
import io
from typing import Dict, List, Optional, Any


def excel_to_structured_data(file_bytes: bytes, file_path: str, sheet_names: Optional[List[str]] = None) -> Dict[str, Any]:
    """
    Convert Excel file to structured review artifacts (CSV/JSON per sheet).
    
    Args:
        file_bytes: Excel file content as bytes
        file_path: Original file path
        sheet_names: Optional list of sheet names to convert (if None, converts all)
    
    Returns:
        Dictionary with converted data:
        {
            "file_path": "...",
            "file_name": "...",
            "sheets": [
                {
                    "sheet_name": "Door Schedule",
                    "format": "csv",  # or "json"
                    "content": "...",  # CSV or JSON string
                    "columns": ["Door No", "Type", "Rating"],
                    "row_count": 25
                }
            ]
        }
    """
    try:
        import openpyxl
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl library required for Excel conversion")
    
    # Write to temp file
    file_ext = os.path.splitext(file_path)[1].lower()
    fd, tmp_path = tempfile.mkstemp(suffix=file_ext)
    
    try:
        with os.fdopen(fd, 'wb') as tmp_file:
            tmp_file.write(file_bytes)
        
        # Load workbook
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        
        # Determine which sheets to convert
        sheets_to_convert = sheet_names if sheet_names else wb.sheetnames
        
        converted_sheets = []
        
        for sheet_name in sheets_to_convert:
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            
            # Extract data
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Filter out completely empty rows
                if any(cell is not None and str(cell).strip() for cell in row):
                    rows.append(row)
            
            if not rows:
                continue
            
            # Get column headers (first row)
            headers = [str(cell) if cell is not None else "" for cell in rows[0]]
            
            # Remove empty trailing columns
            while headers and not headers[-1]:
                headers.pop()
            
            # Convert to CSV format (preferred for tables)
            csv_output = io.StringIO()
            csv_writer = csv.writer(csv_output)
            
            # Write headers
            csv_writer.writerow(headers)
            
            # Write data rows
            for row in rows[1:]:
                # Truncate row to match header length
                row_data = [str(cell) if cell is not None else "" for cell in row[:len(headers)]]
                csv_writer.writerow(row_data)
            
            csv_content = csv_output.getvalue()
            
            converted_sheets.append({
                "sheet_name": sheet_name,
                "format": "csv",
                "content": csv_content,
                "columns": headers,
                "row_count": len(rows) - 1  # Exclude header
            })
        
        wb.close()
        
        return {
            "file_path": file_path,
            "file_name": os.path.basename(file_path),
            "sheets": converted_sheets
        }
    
    finally:
        try:
            os.unlink(tmp_path)
        except:
            pass


def excel_to_review_artifacts(file_bytes: bytes, file_path: str) -> List[Dict[str, Any]]:
    """
    Convert Excel file to review artifacts for GPT vision API.
    
    Returns list of content items:
    [
        {"type": "text", "text": "Sheet: Door Schedule\nCSV:\n..."},
        ...
    ]
    """
    structured_data = excel_to_structured_data(file_bytes, file_path)
    
    artifacts = []
    
    for sheet_data in structured_data["sheets"]:
        sheet_text = f"""File: {structured_data['file_name']}
Path: {structured_data['file_path']}
Sheet: {sheet_data['sheet_name']}
Columns: {', '.join(sheet_data['columns'])}
Row Count: {sheet_data['row_count']}

CSV Data:
{sheet_data['content']}

Review this schedule for constructability issues. Check for missing items, inconsistencies, and coordination problems."""
        
        artifacts.append({
            "type": "text",
            "text": sheet_text
        })
    
    return artifacts


def dwg_to_pdf_instructions() -> str:
    """
    Returns instructions for DWG to PDF conversion.
    Since DWG conversion requires external tools (AutoCAD, DWG TrueView, etc.),
    this provides guidance for manual or automated conversion.
    
    Returns:
        Instructions string for conversion process
    """
    return """
DWG files require conversion to PDF before review. 

Conversion options:
1. Use AutoCAD batch plot to export layouts to PDF
2. Use DWG TrueView to convert DWG to PDF
3. Use headless CAD service if available
4. Use online conversion service

Once converted, the PDF can be processed normally.
"""


def prepare_file_for_review(file_path: str, file_bytes: bytes, doc_type: str) -> List[Dict[str, Any]]:
    """
    Prepare any file type for GPT review by converting to appropriate format.
    
    Args:
        file_path: Path to the file
        file_bytes: File content as bytes
        doc_type: File type ("pdf", "xls", "xlsx", "dwg", etc.)
    
    Returns:
        List of content items for vision API
    """
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if doc_type in ["xls", "xlsx"]:
        # Convert Excel to structured text
        return excel_to_review_artifacts(file_bytes, file_path)
    
    elif doc_type == "dwg":
        # DWG needs external conversion - return instructions
        return [{
            "type": "text",
            "text": f"File: {os.path.basename(file_path)}\nPath: {file_path}\n\n{dwg_to_pdf_instructions()}\n\nNote: This DWG file needs to be converted to PDF before review."
        }]
    
    elif doc_type in ["txt", "csv", "json", "xml"]:
        # Text files - decode and return as text
        try:
            text_content = file_bytes.decode('utf-8')
        except:
            try:
                text_content = file_bytes.decode('latin-1')
            except:
                text_content = f"File content (binary, {len(file_bytes)} bytes)"
        
        return [{
            "type": "text",
            "text": f"File: {os.path.basename(file_path)}\nPath: {file_path}\n\n{text_content}"
        }]
    
    else:
        # Unknown file type - return as base64 for now
        file_base64 = base64.b64encode(file_bytes).decode('utf-8')
        return [{
            "type": "text",
            "text": f"File: {os.path.basename(file_path)}\nPath: {file_path}\nType: {doc_type}\n\nFile content (base64 encoded, {len(file_bytes)} bytes). This file type may need special handling."
        }]

